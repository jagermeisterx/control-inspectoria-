import io
from datetime import date, timedelta
from collections import Counter

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Count, Q
from django.utils import timezone

import openpyxl
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from .models import Alumno, Retiro, Atraso, ControlUniforme, Celular, VisitaApoderado
from .forms import (
    AlumnoForm, RetiroForm, AtrasoForm, ControlUniformeForm,
    CelularForm, VisitaApoderadoForm, ImportAlumnosForm,
)


# ── Dashboard ──
@login_required
def dashboard(request):
    hoy = date.today()
    mes = hoy.month
    anio = hoy.year

    retiros_mes = Retiro.objects.filter(fecha__month=mes, fecha__year=anio).count()
    atrasos_mes = Atraso.objects.filter(fecha__month=mes, fecha__year=anio).count()
    uniformes_mes = ControlUniforme.objects.filter(fecha__month=mes, fecha__year=anio).count()
    celulares_mes = Celular.objects.filter(fecha__month=mes, fecha__year=anio).count()
    visitas_mes = VisitaApoderado.objects.filter(fecha__month=mes, fecha__year=anio).count()

    retiros_hoy = Retiro.objects.filter(fecha=hoy).count()
    atrasos_hoy = Atraso.objects.filter(fecha=hoy).count()

    # Top atrasos por alumno (mes actual)
    top_atrasos = (
        Atraso.objects.filter(fecha__month=mes, fecha__year=anio)
        .values("alumno__nombre", "alumno__apellido", "alumno__curso")
        .annotate(total=Count("id"))
        .order_by("-total")[:10]
    )

    # Atrasos por curso
    atrasos_por_curso = (
        Atraso.objects.filter(fecha__month=mes, fecha__year=anio)
        .values("alumno__curso")
        .annotate(total=Count("id"))
        .order_by("-total")
    )

    # Retiros por motivo
    retiros_por_motivo = (
        Retiro.objects.filter(fecha__month=mes, fecha__year=anio)
        .values("motivo")
        .annotate(total=Count("id"))
        .order_by("-total")
    )

    # Últimos 10 registros de cada tipo
    ultimos_retiros = Retiro.objects.select_related("alumno")[:5]
    ultimos_atrasos = Atraso.objects.select_related("alumno")[:5]

    ctx = {
        "retiros_mes": retiros_mes,
        "atrasos_mes": atrasos_mes,
        "uniformes_mes": uniformes_mes,
        "celulares_mes": celulares_mes,
        "visitas_mes": visitas_mes,
        "retiros_hoy": retiros_hoy,
        "atrasos_hoy": atrasos_hoy,
        "top_atrasos": top_atrasos,
        "atrasos_por_curso": atrasos_por_curso,
        "retiros_por_motivo": retiros_por_motivo,
        "ultimos_retiros": ultimos_retiros,
        "ultimos_atrasos": ultimos_atrasos,
        "total_alumnos": Alumno.objects.filter(activo=True).count(),
        "mes_nombre": hoy.strftime("%B %Y").capitalize(),
    }
    return render(request, "core/dashboard.html", ctx)


# ── Generic list+create pattern ──
def _list_create(request, model, form_class, template, extra_context=None):
    form = form_class(request.POST or None)
    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        if hasattr(obj, "registrado_por"):
            obj.registrado_por = request.user
        obj.save()
        messages.success(request, "Registro guardado correctamente.")
        return redirect(request.path)

    qs = model.objects.select_related("alumno") if hasattr(model, "alumno") else model.objects.all()

    # Filtro por fecha
    fecha_desde = request.GET.get("fecha_desde")
    fecha_hasta = request.GET.get("fecha_hasta")
    buscar = request.GET.get("buscar", "")

    if fecha_desde:
        qs = qs.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha__lte=fecha_hasta)
    if buscar and hasattr(model, "alumno"):
        qs = qs.filter(
            Q(alumno__nombre__icontains=buscar) | Q(alumno__apellido__icontains=buscar)
        )

    ctx = {"form": form, "registros": qs[:200], "buscar": buscar, "fecha_desde": fecha_desde or "", "fecha_hasta": fecha_hasta or ""}
    if extra_context:
        ctx.update(extra_context)
    return render(request, template, ctx)


@login_required
def retiros(request):
    return _list_create(request, Retiro, RetiroForm, "core/retiros.html")

@login_required
def atrasos(request):
    return _list_create(request, Atraso, AtrasoForm, "core/atrasos.html")

@login_required
def uniformes(request):
    return _list_create(request, ControlUniforme, ControlUniformeForm, "core/uniformes.html")

@login_required
def celulares(request):
    return _list_create(request, Celular, CelularForm, "core/celulares.html")

@login_required
def visitas(request):
    form = VisitaApoderadoForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.registrado_por = request.user
        obj.save()
        messages.success(request, "Visita registrada correctamente.")
        return redirect("visitas")

    qs = VisitaApoderado.objects.all()
    fecha_desde = request.GET.get("fecha_desde")
    fecha_hasta = request.GET.get("fecha_hasta")
    if fecha_desde:
        qs = qs.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha__lte=fecha_hasta)

    return render(request, "core/visitas.html", {"form": form, "registros": qs[:200], "fecha_desde": fecha_desde or "", "fecha_hasta": fecha_hasta or ""})


# ── Delete ──
@login_required
def eliminar_registro(request, modelo, pk):
    modelos = {
        "retiro": Retiro,
        "atraso": Atraso,
        "uniforme": ControlUniforme,
        "celular": Celular,
        "visita": VisitaApoderado,
    }
    redir = {
        "retiro": "retiros",
        "atraso": "atrasos",
        "uniforme": "uniformes",
        "celular": "celulares",
        "visita": "visitas",
    }
    model = modelos.get(modelo)
    if not model:
        messages.error(request, "Tipo de registro inválido.")
        return redirect("dashboard")

    obj = get_object_or_404(model, pk=pk)
    if request.method == "POST":
        obj.delete()
        messages.success(request, "Registro eliminado.")
    return redirect(redir[modelo])


# ── Alumnos ──
@login_required
def alumnos(request):
    form = AlumnoForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Alumno agregado.")
        return redirect("alumnos")

    buscar = request.GET.get("buscar", "")
    curso_filter = request.GET.get("curso", "")
    qs = Alumno.objects.filter(activo=True)
    if buscar:
        qs = qs.filter(Q(nombre__icontains=buscar) | Q(apellido__icontains=buscar))
    if curso_filter:
        qs = qs.filter(curso=curso_filter)

    from .models import CURSOS
    return render(request, "core/alumnos.html", {
        "form": form, "alumnos": qs[:300], "buscar": buscar,
        "curso_filter": curso_filter, "cursos": CURSOS,
    })


@login_required
def importar_alumnos(request):
    if request.method == "POST":
        form = ImportAlumnosForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES["archivo"]
            try:
                wb = openpyxl.load_workbook(archivo, read_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(min_row=2, values_only=True))
                count = 0
                for row in rows:
                    if not row or not row[0]:
                        continue
                    nombre = str(row[0]).strip().upper() if row[0] else ""
                    apellido = str(row[1]).strip().upper() if len(row) > 1 and row[1] else ""
                    curso = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                    rut = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                    apod_nombre = str(row[4]).strip().upper() if len(row) > 4 and row[4] else ""
                    apod_tel = str(row[5]).strip() if len(row) > 5 and row[5] else ""

                    if nombre and apellido:
                        Alumno.objects.update_or_create(
                            nombre=nombre, apellido=apellido, anio=date.today().year,
                            defaults={"curso": curso, "rut": rut, "apoderado_nombre": apod_nombre, "apoderado_telefono": apod_tel},
                        )
                        count += 1
                messages.success(request, f"Se importaron {count} alumnos correctamente.")
                return redirect("alumnos")
            except Exception as e:
                messages.error(request, f"Error al procesar el archivo: {e}")
    else:
        form = ImportAlumnosForm()
    return render(request, "core/importar_alumnos.html", {"form": form})


# ── Reportes ──
@login_required
def reportes(request):
    return render(request, "core/reportes.html", {
        "alumnos": Alumno.objects.filter(activo=True),
        "cursos": Alumno.objects.filter(activo=True).exclude(curso="").values_list("curso", flat=True).distinct().order_by("curso"),
    })


@login_required
def reporte_alumno(request, pk):
    alumno = get_object_or_404(Alumno, pk=pk)
    retiros = Retiro.objects.filter(alumno=alumno)
    atrs = Atraso.objects.filter(alumno=alumno)
    unifs = ControlUniforme.objects.filter(alumno=alumno)
    cels = Celular.objects.filter(alumno=alumno)

    ctx = {
        "alumno": alumno,
        "retiros": retiros,
        "atrasos": atrs,
        "uniformes": unifs,
        "celulares": cels,
        "total_retiros": retiros.count(),
        "total_atrasos": atrs.count(),
        "total_uniformes": unifs.count(),
        "total_celulares": cels.count(),
    }
    return render(request, "core/reporte_alumno.html", ctx)


@login_required
def reporte_curso(request, curso):
    alumnos_curso = Alumno.objects.filter(curso=curso, activo=True)
    datos = []
    for al in alumnos_curso:
        datos.append({
            "alumno": al,
            "retiros": al.retiros.count(),
            "atrasos": al.atrasos.count(),
            "uniformes": al.uniformes.count(),
            "celulares": al.celulares.count(),
        })
    datos.sort(key=lambda x: x["atrasos"] + x["retiros"], reverse=True)
    return render(request, "core/reporte_curso.html", {"curso": curso, "datos": datos})


# ── Exportar PDF ──
@login_required
def exportar_pdf_alumno(request, pk):
    alumno = get_object_or_404(Alumno, pk=pk)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=20*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"Reporte de Inspectoría — {alumno.nombre_completo}", styles["Title"]))
    elements.append(Paragraph(f"Curso: {alumno.curso} | RUT: {alumno.rut or 'N/A'}", styles["Normal"]))
    elements.append(Paragraph(f"Apoderado: {alumno.apoderado_nombre or 'N/A'} | Tel: {alumno.apoderado_telefono or 'N/A'}", styles["Normal"]))
    elements.append(Spacer(1, 10*mm))

    def make_table(title, headers, rows):
        elements.append(Paragraph(title, styles["Heading2"]))
        if not rows:
            elements.append(Paragraph("Sin registros.", styles["Normal"]))
        else:
            data = [headers] + rows
            t = Table(data, repeatRows=1)
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4f46e5")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]))
            elements.append(t)
        elements.append(Spacer(1, 6*mm))

    retiros = Retiro.objects.filter(alumno=alumno)
    make_table(f"Retiros ({retiros.count()})",
        ["Fecha", "Hora", "Motivo", "Retira", "Observación"],
        [[str(r.fecha), str(r.hora)[:5], r.motivo, r.persona_retira, r.observacion[:50]] for r in retiros])

    atrs = Atraso.objects.filter(alumno=alumno)
    make_table(f"Atrasos ({atrs.count()})",
        ["Fecha", "Hora", "Tipo", "Lugar"],
        [[str(a.fecha), str(a.hora)[:5], a.tipo, a.lugar] for a in atrs])

    unifs = ControlUniforme.objects.filter(alumno=alumno)
    make_table(f"Uniformes ({unifs.count()})",
        ["Fecha", "Falta", "Comprado", "Detalle"],
        [[str(u.fecha), u.falta, "Sí" if u.tiene_uniforme_comprado else "No", u.detalle[:50]] for u in unifs])

    cels = Celular.objects.filter(alumno=alumno)
    make_table(f"Celulares ({cels.count()})",
        ["Fecha", "Lugar", "Retiro", "Aviso"],
        [[str(c.fecha), c.lugar_entregado, c.retiro, "Sí" if c.aviso_apoderado else "No"] for c in cels])

    doc.build(elements)
    buf.seek(0)
    resp = HttpResponse(buf, content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="reporte_{alumno.apellido}_{alumno.nombre}.pdf"'
    return resp


@login_required
def exportar_pdf_curso(request, curso):
    alumnos_curso = Alumno.objects.filter(curso=curso, activo=True)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter), topMargin=15*mm, bottomMargin=10*mm)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"Reporte de Inspectoría — {curso}", styles["Title"]))
    elements.append(Paragraph(f"Generado: {date.today().strftime('%d/%m/%Y')}", styles["Normal"]))
    elements.append(Spacer(1, 8*mm))

    headers = ["Alumno/a", "Retiros", "Atrasos", "Uniformes", "Celulares", "Total"]
    rows = []
    for al in alumnos_curso:
        r = al.retiros.count()
        a = al.atrasos.count()
        u = al.uniformes.count()
        c = al.celulares.count()
        rows.append([al.nombre_completo, str(r), str(a), str(u), str(c), str(r+a+u+c)])
    rows.sort(key=lambda x: int(x[5]), reverse=True)

    data = [headers] + rows
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4f46e5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
    ]))
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    resp = HttpResponse(buf, content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="reporte_{curso.replace(" ","_")}.pdf"'
    return resp


# ── Exportar Excel ──
@login_required
def exportar_excel_alumno(request, pk):
    alumno = get_object_or_404(Alumno, pk=pk)
    wb = openpyxl.Workbook()

    # Retiros
    ws = wb.active
    ws.title = "Retiros"
    ws.append(["Fecha", "Hora", "Motivo", "Persona que retira", "RUT", "Observación"])
    for r in Retiro.objects.filter(alumno=alumno):
        ws.append([r.fecha, r.hora.strftime("%H:%M"), r.motivo, r.persona_retira, r.rut_retira, r.observacion])

    # Atrasos
    ws2 = wb.create_sheet("Atrasos")
    ws2.append(["Fecha", "Hora", "Tipo", "Lugar", "Observación"])
    for a in Atraso.objects.filter(alumno=alumno):
        ws2.append([a.fecha, a.hora.strftime("%H:%M"), a.tipo, a.lugar, a.observacion])

    # Uniformes
    ws3 = wb.create_sheet("Uniformes")
    ws3.append(["Fecha", "Falta", "Comprado", "Detalle", "Llamado", "Determinación"])
    for u in ControlUniforme.objects.filter(alumno=alumno):
        ws3.append([u.fecha, u.falta, "Sí" if u.tiene_uniforme_comprado else "No", u.detalle, "Sí" if u.llamado else "No", u.determinacion])

    # Celulares
    ws4 = wb.create_sheet("Celulares")
    ws4.append(["Fecha", "Lugar", "Retiro", "Aviso apoderado"])
    for c in Celular.objects.filter(alumno=alumno):
        ws4.append([c.fecha, c.lugar_entregado, c.retiro, "Sí" if c.aviso_apoderado else "No"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="reporte_{alumno.apellido}_{alumno.nombre}.xlsx"'
    return resp


@login_required
def exportar_excel_curso(request, curso):
    alumnos_curso = Alumno.objects.filter(curso=curso, activo=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen"
    ws.append(["Alumno/a", "Retiros", "Atrasos", "Uniformes", "Celulares", "Total"])
    for al in alumnos_curso:
        r = al.retiros.count()
        a = al.atrasos.count()
        u = al.uniformes.count()
        c = al.celulares.count()
        ws.append([al.nombre_completo, r, a, u, c, r+a+u+c])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="reporte_{curso.replace(" ","_")}.xlsx"'
    return resp


# ── API: buscar alumnos (para autocompletar) ──
@login_required
def api_buscar_alumnos(request):
    from django.http import JsonResponse
    q = request.GET.get("q", "")
    if len(q) < 2:
        return JsonResponse([], safe=False)
    alumnos = Alumno.objects.filter(
        Q(nombre__icontains=q) | Q(apellido__icontains=q), activo=True
    )[:15]
    data = [{"id": a.id, "text": f"{a.nombre_completo} ({a.curso})"} for a in alumnos]
    return JsonResponse(data, safe=False)


# ── Cargar datos históricos desde Excel ──
@login_required
def cargar_historico(request):
    if not request.user.is_superuser:
        messages.error(request, "Solo administradores pueden cargar datos históricos.")
        return redirect("dashboard")

    if request.method == "POST" and request.FILES.get("archivo"):
        archivo = request.FILES["archivo"]
        try:
            from datetime import datetime as dt, time as t
            wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
            resultados = []

            def parse_date(val):
                if isinstance(val, date):
                    return val
                if hasattr(val, 'date'):
                    return val.date()
                if isinstance(val, str) and val:
                    try:
                        return dt.strptime(val, "%Y-%m-%d").date()
                    except ValueError:
                        return None
                return None

            def parse_time(val):
                if isinstance(val, t):
                    return val
                if hasattr(val, 'time'):
                    return val.time()
                if isinstance(val, str) and val:
                    try:
                        return dt.strptime(val, "%H:%M").time()
                    except ValueError:
                        return t(8, 0)
                return t(8, 0)

            def get_alumno(nombre, apellido, curso):
                nombre = (nombre or "").strip().upper()
                apellido = (apellido or "").strip().upper()
                if not nombre or not apellido:
                    return None
                al, _ = Alumno.objects.get_or_create(
                    nombre=nombre, apellido=apellido, anio=date.today().year,
                    defaults={"curso": curso or ""},
                )
                return al

            # ── Retiros ──
            if "Retiros" in wb.sheetnames:
                count = 0
                for row in wb["Retiros"].iter_rows(min_row=2, values_only=True):
                    fecha = parse_date(row[0])
                    if not fecha or not row[1] or not row[2]:
                        continue
                    al = get_alumno(row[1], row[2], row[3])
                    if not al:
                        continue
                    Retiro.objects.create(
                        alumno=al, fecha=fecha, hora=parse_time(row[5]),
                        motivo=str(row[4] or "OTRO"),
                        persona_retira=str(row[6] or ""),
                        rut_retira=str(row[7] or ""),
                        registrado_por=request.user,
                    )
                    count += 1
                resultados.append(f"✓ {count} retiros")

            # ── Atrasos ──
            if "Atrasos" in wb.sheetnames:
                count = 0
                for row in wb["Atrasos"].iter_rows(min_row=2, values_only=True):
                    fecha = parse_date(row[0])
                    if not fecha or not row[1] or not row[2]:
                        continue
                    al = get_alumno(row[1], row[2], row[3])
                    if not al:
                        continue
                    Atraso.objects.create(
                        alumno=al, fecha=fecha, hora=parse_time(row[4]),
                        tipo=str(row[5] or "LLEGADA"),
                        lugar=str(row[6] or ""),
                        registrado_por=request.user,
                    )
                    count += 1
                resultados.append(f"✓ {count} atrasos")

            # ── Uniformes ──
            if "Uniformes" in wb.sheetnames:
                count = 0
                for row in wb["Uniformes"].iter_rows(min_row=2, values_only=True):
                    fecha = parse_date(row[0])
                    if not fecha or not row[1] or not row[2]:
                        continue
                    al = get_alumno(row[1], row[2], row[3])
                    if not al:
                        continue
                    comprado = str(row[5] or "").upper() in ("SI", "SÍ")
                    llamado = str(row[8] or "").upper() in ("SI", "SÍ")
                    ControlUniforme.objects.create(
                        alumno=al, fecha=fecha,
                        falta=str(row[4] or "SIN UNIFORME"),
                        tiene_uniforme_comprado=comprado,
                        detalle=str(row[6] or ""),
                        contacto_apoderado=str(row[7] or ""),
                        llamado=llamado,
                        registrado_por=request.user,
                    )
                    count += 1
                resultados.append(f"✓ {count} uniformes")

            # ── Celulares ──
            if "Celulares" in wb.sheetnames:
                count = 0
                for row in wb["Celulares"].iter_rows(min_row=2, values_only=True):
                    fecha = parse_date(row[0])
                    if not fecha or not row[1] or not row[2]:
                        continue
                    al = get_alumno(row[1], row[2], row[3])
                    if not al:
                        continue
                    aviso = str(row[6] or "").upper() in ("SI", "SÍ")
                    Celular.objects.create(
                        alumno=al, fecha=fecha,
                        lugar_entregado=str(row[4] or "DIRECCIÓN"),
                        retiro=str(row[5] or "AL FINAL DEL DÍA"),
                        aviso_apoderado=aviso,
                        registrado_por=request.user,
                    )
                    count += 1
                resultados.append(f"✓ {count} celulares")

            # ── Visitas ──
            if "Visitas" in wb.sheetnames:
                count = 0
                for row in wb["Visitas"].iter_rows(min_row=2, values_only=True):
                    fecha = parse_date(row[0])
                    if not fecha:
                        continue
                    hora = parse_time(row[1]) if row[1] else None
                    VisitaApoderado.objects.create(
                        fecha=fecha, hora=hora,
                        destino=str(row[2] or "INSPECTORÍA GENERAL"),
                        funcionario=str(row[3] or ""),
                        registrado_por=request.user,
                    )
                    count += 1
                resultados.append(f"✓ {count} visitas")

            wb.close()
            messages.success(request, f"Carga completa: {' · '.join(resultados)}")
            return redirect("dashboard")

        except Exception as e:
            messages.error(request, f"Error al procesar el archivo: {e}")

    return render(request, "core/cargar_historico.html")
