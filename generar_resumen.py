import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
from datetime import datetime

files = [
    ('MARZO - REPORTE INSPECTORÍA.xlsx', 'MARZO'),
    ('ABRIL - REPORTE INSPECTORÍA.xlsx', 'ABRIL'),
    ('MAYO - REPORTE INSPECTORÍA.xlsx',  'MAYO'),
    ('JUNIO - REPORTE INSPECTORÍA.xlsx', 'JUNIO'),
    ('JULIO - REPORTE INSPECTORÍA.xlsx', 'JULIO'),
]

target = 'AGUSTIN VEGA'

registros = []

for fname, mes in files:
    path = os.path.join('datos', fname)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['ATRASOS  PASES']
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        continue
    header = rows[0]

    idx_fecha = header.index('FECHA')
    idx_nom   = header.index('NOMBRE_ALUMNO')
    idx_ape   = header.index('APELLIDO_ALUMNO')
    idx_curso = header.index('CURSO')
    idx_hora  = header.index('HORA')
    idx_tipo  = header.index('LLEGADA O RECREO')
    idx_lugar = None
    for cand in ('LUGAR', 'RAZÓN'):
        if cand in header:
            idx_lugar = header.index(cand)
            break

    for row in rows[1:]:
        if not row or row[idx_fecha] is None:
            continue
        nom = (str(row[idx_nom]) if row[idx_nom] else '').strip().upper()
        ape = (str(row[idx_ape]) if row[idx_ape] else '').strip().upper()
        if nom == 'AGUSTIN' and ape == 'VEGA':
            fecha = row[idx_fecha]
            meses = {1:'ENERO',2:'FEBRERO',3:'MARZO',4:'ABRIL',5:'MAYO',6:'JUNIO',7:'JULIO',8:'AGOSTO',9:'SEPTIEMBRE',10:'OCTUBRE',11:'NOVIEMBRE',12:'DICIEMBRE'}
            mes_real = meses[fecha.month] if hasattr(fecha, 'month') else mes
            registros.append({
                'MES':       mes_real,
                'FECHA':     fecha,
                'NOMBRE':    'AGUSTIN',
                'APELLIDO':  'VEGA',
                'CURSO':     row[idx_curso],
                'HORA':      row[idx_hora],
                'TIPO':      row[idx_tipo],
                'LUGAR/RAZÓN': row[idx_lugar] if idx_lugar is not None and len(row) > idx_lugar else None,
            })

registros.sort(key=lambda r: (r['FECHA'], r['HORA']))

out = openpyxl.Workbook()
ws = out.active
ws.title = 'ATRASOS AGUSTIN VEGA'

header_out = ['MES', 'FECHA', 'NOMBRE', 'APELLIDO', 'CURSO', 'HORA', 'TIPO', 'LUGAR/RAZÓN']
ws.append(header_out)

header_fill = PatternFill('solid', fgColor='305496')
header_font = Font(bold=True, color='FFFFFF', size=11)
border = Border(
    left=Side(style='thin', color='B4B4B4'),
    right=Side(style='thin', color='B4B4B4'),
    top=Side(style='thin', color='B4B4B4'),
    bottom=Side(style='thin', color='B4B4B4'),
)
for col_idx, _ in enumerate(header_out, start=1):
    c = ws.cell(row=1, column=col_idx)
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border

fill_llegada = PatternFill('solid', fgColor='FCE4D6')
fill_recreo  = PatternFill('solid', fgColor='DDEBF7')

for r in registros:
    ws.append([
        r['MES'],
        r['FECHA'],
        r['NOMBRE'],
        r['APELLIDO'],
        r['CURSO'],
        r['HORA'],
        r['TIPO'],
        r['LUGAR/RAZÓN'],
    ])
    row_idx = ws.max_row
    for col_idx in range(1, len(header_out) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = border
        cell.alignment = Alignment(vertical='center')
        if col_idx == 2 and isinstance(r['FECHA'], datetime):
            cell.number_format = 'DD/MM/YYYY'
        if col_idx == 6 and hasattr(r['HORA'], 'strftime'):
            cell.number_format = 'HH:MM'
    if r['TIPO'] == 'LLEGADA':
        ws.cell(row=row_idx, column=7).fill = fill_llegada
    elif r['TIPO'] == 'RECREO':
        ws.cell(row=row_idx, column=7).fill = fill_recreo

widths = {'A': 10, 'B': 13, 'C': 12, 'D': 12, 'E': 14, 'F': 8, 'G': 18, 'H': 28}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions

last_data_row = ws.max_row
total_row = last_data_row + 2
ws.cell(row=total_row, column=1, value='TOTAL ATRASOS').font = Font(bold=True)
ws.cell(row=total_row, column=2, value=len(registros)).font = Font(bold=True)

ws2 = out.create_sheet('RESUMEN POR MES')
ws2.append(['MES', 'CANTIDAD DE ATRASOS'])
for col_idx in range(1, 3):
    c = ws2.cell(row=1, column=col_idx)
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border

from collections import Counter
conteo = Counter(r['MES'] for r in registros)
orden = ['MARZO', 'ABRIL', 'MAYO', 'JUNIO', 'JULIO']
for m in orden:
    ws2.append([m, conteo.get(m, 0)])
    row_idx = ws2.max_row
    for col_idx in range(1, 3):
        ws2.cell(row=row_idx, column=col_idx).border = border
        ws2.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal='center')

ws2.append(['TOTAL', sum(conteo.values())])
total_idx = ws2.max_row
for col_idx in range(1, 3):
    cell = ws2.cell(row=total_idx, column=col_idx)
    cell.font = Font(bold=True)
    cell.border = border
    cell.fill = PatternFill('solid', fgColor='FFE699')
    cell.alignment = Alignment(horizontal='center')

ws2.column_dimensions['A'].width = 14
ws2.column_dimensions['B'].width = 22

out_path = 'RESUMEN_ATRASOS_AGUSTIN_VEGA.xlsx'
out.save(out_path)
print('OK ->', out_path, '| registros:', len(registros))
for m in orden:
    print(f'  {m}: {conteo.get(m, 0)}')
