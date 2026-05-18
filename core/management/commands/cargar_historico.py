"""
Comando para cargar datos históricos desde datos_historicos.xlsx
Uso: python manage.py cargar_historico datos_historicos.xlsx
"""
from datetime import datetime, time
from django.core.management.base import BaseCommand
from core.models import Alumno, Retiro, Atraso, ControlUniforme, Celular, VisitaApoderado
import openpyxl


class Command(BaseCommand):
    help = "Carga datos históricos desde datos_historicos.xlsx"

    def add_arguments(self, parser):
        parser.add_argument("archivo", type=str, help="Ruta al archivo .xlsx")

    def _get_or_create_alumno(self, nombre, apellido, curso):
        nombre = (nombre or "").strip().upper()
        apellido = (apellido or "").strip().upper()
        if not nombre or not apellido:
            return None
        al, _ = Alumno.objects.get_or_create(
            nombre=nombre, apellido=apellido, anio=2026,
            defaults={"curso": curso or ""},
        )
        if curso and not al.curso:
            al.curso = curso
            al.save(update_fields=["curso"])
        return al

    def _parse_date(self, val):
        if isinstance(val, datetime):
            return val.date()
        if isinstance(val, str) and val:
            try:
                return datetime.strptime(val, "%Y-%m-%d").date()
            except ValueError:
                return None
        return None

    def _parse_time(self, val):
        if isinstance(val, time):
            return val
        if isinstance(val, datetime):
            return val.time()
        if isinstance(val, str) and val:
            try:
                return datetime.strptime(val, "%H:%M").time()
            except ValueError:
                return time(8, 0)
        return time(8, 0)

    def handle(self, *args, **options):
        archivo = options["archivo"]
        wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)

        # ── Retiros ──
        if "Retiros" in wb.sheetnames:
            ws = wb["Retiros"]
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                fecha = self._parse_date(row[0])
                if not fecha or not row[1] or not row[2]:
                    continue
                al = self._get_or_create_alumno(row[1], row[2], row[3])
                if not al:
                    continue
                Retiro.objects.create(
                    alumno=al, fecha=fecha, hora=self._parse_time(row[5]),
                    motivo=str(row[4] or "OTRO"),
                    persona_retira=str(row[6] or ""),
                    rut_retira=str(row[7] or ""),
                )
                count += 1
            self.stdout.write(self.style.SUCCESS(f"Retiros cargados: {count}"))

        # ── Atrasos ──
        if "Atrasos" in wb.sheetnames:
            ws = wb["Atrasos"]
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                fecha = self._parse_date(row[0])
                if not fecha or not row[1] or not row[2]:
                    continue
                al = self._get_or_create_alumno(row[1], row[2], row[3])
                if not al:
                    continue
                Atraso.objects.create(
                    alumno=al, fecha=fecha, hora=self._parse_time(row[4]),
                    tipo=str(row[5] or "LLEGADA"),
                    lugar=str(row[6] or ""),
                )
                count += 1
            self.stdout.write(self.style.SUCCESS(f"Atrasos cargados: {count}"))

        # ── Uniformes ──
        if "Uniformes" in wb.sheetnames:
            ws = wb["Uniformes"]
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                fecha = self._parse_date(row[0])
                if not fecha or not row[1] or not row[2]:
                    continue
                al = self._get_or_create_alumno(row[1], row[2], row[3])
                if not al:
                    continue
                comprado = str(row[5] or "").upper() in ("SI", "SÍ")
                llamado = str(row[8] or "").upper() in ("SI", "SÍ")
                ControlUniforme.objects.create(
                    alumno=al, fecha=fecha,
                    falta=str(row[4] or "SIN UNIFORME"),
                    tiene_uniforme_comprado=comprado,
                    detalle=str(row[6] or ""),
                    contacto_apoderado=str(row[7] or ""),
                    llamado=llamado,
                )
                count += 1
            self.stdout.write(self.style.SUCCESS(f"Uniformes cargados: {count}"))

        # ── Celulares ──
        if "Celulares" in wb.sheetnames:
            ws = wb["Celulares"]
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                fecha = self._parse_date(row[0])
                if not fecha or not row[1] or not row[2]:
                    continue
                al = self._get_or_create_alumno(row[1], row[2], row[3])
                if not al:
                    continue
                aviso = str(row[6] or "").upper() in ("SI", "SÍ")
                Celular.objects.create(
                    alumno=al, fecha=fecha,
                    lugar_entregado=str(row[4] or "DIRECCIÓN"),
                    retiro=str(row[5] or "AL FINAL DEL DÍA"),
                    aviso_apoderado=aviso,
                )
                count += 1
            self.stdout.write(self.style.SUCCESS(f"Celulares cargados: {count}"))

        # ── Visitas ──
        if "Visitas" in wb.sheetnames:
            ws = wb["Visitas"]
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                fecha = self._parse_date(row[0])
                if not fecha:
                    continue
                hora = self._parse_time(row[1]) if row[1] else None
                VisitaApoderado.objects.create(
                    fecha=fecha,
                    hora=hora,
                    destino=str(row[2] or "INSPECTORÍA GENERAL"),
                    funcionario=str(row[3] or ""),
                )
                count += 1
            self.stdout.write(self.style.SUCCESS(f"Visitas cargadas: {count}"))

        wb.close()
        self.stdout.write(self.style.SUCCESS("\n¡Carga completa!"))
