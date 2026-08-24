import io
import calendar
from datetime import date, timedelta
from collections import Counter

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth import update_session_auth_hash
from django.contrib import messages
from django.http import HttpResponse
from django.db import transaction
from django.db.models import Count, Q
from django.utils import timezone

import openpyxl
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from .models import Alumno, Retiro, Atraso, ControlUniforme, Celular, VisitaApoderado, LlamadaApoderado
from .forms import (
    AlumnoForm, RetiroForm, AtrasoForm, ControlUniformeForm,
    CelularForm, VisitaApoderadoForm, ImportAlumnosForm,
    LlamadaApoderadoForm,
    UsuarioForm, UsuarioCrearForm, UsuarioPasswordForm,
)
from .roles import es_admin, rol_requerido, tiene_rol, solo_admin, INSPECTOR_GENERAL, INSPECTOR, PROFESOR, DIRECTOR
from .cursos_norm import normalizar_curso


# ── Errores HTTP ──
def error_404(request, exception=None):
    return render(request, "404.html", status=404)


# ── Dashboard ──
@login_required
def dashboard(request):
    if not es_admin(request.user) and tiene_rol(request.user, INSPECTOR):
        return redirect("atrasos")
    if not es_admin(request.user) and tiene_rol(request.user, PROFESOR):
        return redirect("reportes")
    if not es_admin(request.user) and tiene_rol(request.user, DIRECTOR):
        return redirect("dashboard_director")

    hoy = date.today()
    mes = hoy.month
    anio = hoy.year

    retiros_mes = Retiro.objects.filter(fecha__month=mes, fecha__year=anio).count()
    atrasos_mes = Atraso.objects.filter(fecha__month=mes, fecha__year=anio).count()
    uniformes_mes = ControlUniforme.objects.filter(fecha__month=mes, fecha__year=anio).count()
    celulares_mes = Celular.objects.filter(fecha__month=mes, fecha__year=anio).count()
    visitas_mes = VisitaApoderado.objects.filter(fecha__month=mes, fecha__year=anio).count()

    retiros_hoy = Retiro.objects.filter(fecha=hoy).count()
    atrasos_hoy = Atraso.objects.filter(fecha=hoy).count()

    # Top atrasos por alumno (mes actual)
    top_atrasos = (
        Atraso.objects.filter(fecha__month=mes, fecha__year=anio)
        .values("alumno__nombre", "alumno__apellido", "alumno__curso")
        .annotate(total=Count("id"))
        .order_by("-total")[:10]
    )

    # Atrasos por curso
    atrasos_por_curso = (
        Atraso.objects.filter(fecha__month=mes, fecha__year=anio)
        .values("alumno__curso")
        .annotate(total=Count("id"))
        .order_by("-total")
    )

    # Retiros por motivo
    retiros_por_motivo = (
        Retiro.objects.filter(fecha__month=mes, fecha__year=anio)
        .values("motivo")
        .annotate(total=Count("id"))
        .order_by("-total")
    )

    # Últimos 10 registros de cada tipo
    ultimos_retiros = Retiro.objects.select_related("alumno")[:5]
    ultimos_atrasos = Atraso.objects.select_related("alumno")[:5]

    ctx = {
        "retiros_mes": retiros_mes,
        "atrasos_mes": atrasos_mes,
        "uniformes_mes": uniformes_mes,
        "celulares_mes": celulares_mes,
        "visitas_mes": visitas_mes,
        "retiros_hoy": retiros_hoy,
        "atrasos_hoy": atrasos_hoy,
        "top_atrasos": top_atrasos,
        "atrasos_por_curso": atrasos_por_curso,
        "retiros_por_motivo": retiros_por_motivo,
        "ultimos_retiros": ultimos_retiros,
        "ultimos_atrasos": ultimos_atrasos,
        "total_alumnos": Alumno.objects.filter(activo=True).count(),
        "mes_nombre": hoy.strftime("%B %Y").capitalize(),
        "hoy_iso": hoy.isoformat(),
        "mes_inicio_iso": date(anio, mes, 1).isoformat(),
        "mes_fin_iso": date(anio, mes, calendar.monthrange(anio, mes)[1]).isoformat(),
    }
    return render(request, "core/dashboard.html", ctx)


# ── Generic list+create pattern ──
def _list_create(request, model, form_class, template, extra_context=None):
    form = form_class(request.POST or None)
    if request.method == "POST":
        alumno_id = request.POST.get("alumno_id")
        if form.is_valid():
            obj = form.save(commit=False)
            # Asignar alumno desde el campo oculto del autocompletado
            if hasattr(obj, "alumno_id") and alumno_id:
                obj.alumno_id = int(alumno_id)
            if hasattr(obj, "registrado_por"):
                obj.registrado_por = request.user
            # Auto-llenar motivo para atrasos segun es_campo
            if isinstance(obj, Atraso) and not obj.motivo:
                try:
                    alumno = Alumno.objects.get(pk=obj.alumno_id)
                    obj.motivo = "CAMPO" if alumno.es_campo else "ATRASO"
                except Alumno.DoesNotExist:
                    obj.motivo = "ATRASO"
            try:
                obj.save()
                messages.success(request, "Registro guardado correctamente.")
                return redirect(request.path)
            except Exception as e:
                messages.error(request, f"Error: {e}")

    qs = model.objects.select_related("alumno") if hasattr(model, "alumno") else model.objects.all()

    # Filtro por fecha
    fecha_desde = request.GET.get("fecha_desde")
    fecha_hasta = request.GET.get("fecha_hasta")
    buscar = request.GET.get("buscar", "")

    if fecha_desde:
        qs = qs.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha__lte=fecha_hasta)
    if buscar and hasattr(model, "alumno"):
        qs = qs.filter(
            Q(alumno__nombre__icontains=buscar) | Q(alumno__apellido__icontains=buscar)
        )

    ctx = {"form": form, "registros": qs[:200], "buscar": buscar, "fecha_desde": fecha_desde or "", "fecha_hasta": fecha_hasta or ""}
    if extra_context:
        ctx.update(extra_context)
    return render(request, template, ctx)


@rol_requerido(INSPECTOR_GENERAL, INSPECTOR)
def retiros(request):
    return _list_create(request, Retiro, RetiroForm, "core/retiros.html")

@rol_requerido(INSPECTOR_GENERAL, INSPECTOR)
def atrasos(request):
    return _list_create(request, Atraso, AtrasoForm, "core/atrasos.html")

@rol_requerido(INSPECTOR_GENERAL, INSPECTOR)
def uniformes(request):
    return _list_create(request, ControlUniforme, ControlUniformeForm, "core/uniformes.html")

@rol_requerido(INSPECTOR_GENERAL, INSPECTOR, PROFESOR)
def celulares(request):
    conteo = dict(
        Celular.objects.values_list("alumno_id").annotate(total=Count("id"))
    )
    return _list_create(
        request, Celular, CelularForm, "core/celulares.html",
        extra_context={"conteo_celulares": conteo},
    )

@rol_requerido(INSPECTOR_GENERAL, INSPECTOR)
def visitas(request):
    form = VisitaApoderadoForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.registrado_por = request.user
        obj.save()
        messages.success(request, "Visita registrada correctamente.")
        return redirect("visitas")

    qs = VisitaApoderado.objects.all()
    fecha_desde = request.GET.get("fecha_desde")
    fecha_hasta = request.GET.get("fecha_hasta")
    if fecha_desde:
        qs = qs.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha__lte=fecha_hasta)

    return render(request, "core/visitas.html", {"form": form, "registros": qs[:200], "fecha_desde": fecha_desde or "", "fecha_hasta": fecha_hasta or ""})


# ── Delete ──
@rol_requerido(INSPECTOR_GENERAL)
def eliminar_registro(request, modelo, pk):
    modelos = {
        "retiro": Retiro,
        "atraso": Atraso,
        "uniforme": ControlUniforme,
        "celular": Celular,
        "visita": VisitaApoderado,
    }
    redir = {
        "retiro": "retiros",
        "atraso": "atrasos",
        "uniforme": "uniformes",
        "celular": "celulares",
        "visita": "visitas",
    }
    model = modelos.get(modelo)
    if not model:
        messages.error(request, "Tipo de registro inválido.")
        return redirect("dashboard")

    obj = get_object_or_404(model, pk=pk)
    if request.method == "POST":
        obj.delete()
        messages.success(request, "Registro eliminado.")
    return redirect(redir[modelo])


# ── Alumnos ──
@rol_requerido(INSPECTOR_GENERAL)
def alumnos(request):
    form = AlumnoForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Alumno agregado.")
        return redirect("alumnos")

    buscar = request.GET.get("buscar", "")
    curso_filter = request.GET.get("curso", "")
    qs = Alumno.objects.filter(activo=True)
    if buscar:
        qs = qs.filter(Q(nombre__icontains=buscar) | Q(apellido__icontains=buscar))
    if curso_filter:
        qs = qs.filter(curso=curso_filter)

    from .models import CURSOS
    return render(request, "core/alumnos.html", {
        "form": form, "alumnos": qs[:300], "buscar": buscar,
        "curso_filter": curso_filter, "cursos": CURSOS,
    })


@rol_requerido(INSPECTOR_GENERAL)
def editar_alumno(request, pk):
    alumno = get_object_or_404(Alumno, pk=pk)
    if request.method != "POST":
        return redirect("alumnos")
    form = AlumnoForm(request.POST, instance=alumno)
    if form.is_valid():
        form.save()
        messages.success(request, f"Datos de {alumno.nombre_completo} actualizados.")
    else:
        errores = " ".join(f"{field}: {' '.join(errs)}" for field, errs in form.errors.items())
        messages.error(request, f"No se pudo actualizar el alumno. {errores}")
    return redirect("alumnos")


@rol_requerido(INSPECTOR_GENERAL, DIRECTOR, PROFESOR)
def toggle_campo(request, pk):
    if request.method != "POST":
        return redirect("alumnos")
    alumno = get_object_or_404(Alumno, pk=pk)
    alumno.es_campo = not alumno.es_campo
    alumno.save(update_fields=["es_campo"])
    estado = "marcado como Campo" if alumno.es_campo else "desmarcado como Campo"
    messages.success(request, f"{alumno.nombre_completo} {estado}.")
    return redirect("alumnos")


@rol_requerido(INSPECTOR_GENERAL)
def importar_alumnos(request):
    if request.method == "POST":
        form = ImportAlumnosForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES["archivo"]
            try:
                anio = date.today().year
                wb = openpyxl.load_workbook(archivo, read_only=True)
                try:
                    ws = wb.active
                    existentes = {
                        (al.nombre, al.apellido, al.anio): al
                        for al in Alumno.objects.filter(anio=anio)
                    }
                    a_crear = []
                    a_actualizar = []
                    count = 0
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row or not row[0]:
                            continue
                        nombre = str(row[0]).strip().upper() if row[0] else ""
                        apellido = str(row[1]).strip().upper() if len(row) > 1 and row[1] else ""
                        curso_raw = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                        curso = normalizar_curso(curso_raw) or curso_raw
                        rut = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                        apod_nombre = str(row[4]).strip().upper() if len(row) > 4 and row[4] else ""
                        apod_tel = str(row[5]).strip() if len(row) > 5 and row[5] else ""

                        if not nombre or not apellido:
                            continue
                        clave = (nombre, apellido, anio)
                        if clave in existentes:
                            al = existentes[clave]
                            al.curso = curso
                            al.rut = rut
                            al.apoderado_nombre = apod_nombre
                            al.apoderado_telefono = apod_tel
                            a_actualizar.append(al)
                        else:
                            a_crear.append(Alumno(
                                nombre=nombre, apellido=apellido, curso=curso,
                                rut=rut, apoderado_nombre=apod_nombre,
                                apoderado_telefono=apod_tel, anio=anio,
                            ))
                        count += 1
                    with transaction.atomic():
                        Alumno.objects.bulk_create(a_crear, batch_size=500)
                        Alumno.objects.bulk_update(
                            a_actualizar,
                            ["curso", "rut", "apoderado_nombre", "apoderado_telefono"],
                            batch_size=500,
                        )
                    messages.success(request, f"Se importaron {count} alumnos correctamente.")
                    return redirect("alumnos")
                finally:
                    wb.close()
            except Exception as e:
                messages.error(request, f"Error al procesar el archivo: {e}")
    else:
        form = ImportAlumnosForm()
    return render(request, "core/importar_alumnos.html", {"form": form})


# ── Reportes ──
@rol_requerido(INSPECTOR_GENERAL, PROFESOR, DIRECTOR)
def reportes(request):
    hoy = date.today()
    return render(request, "core/reportes.html", {
        "alumnos": Alumno.objects.filter(activo=True),
        "cursos": Alumno.objects.filter(activo=True).exclude(curso="").values_list("curso", flat=True).distinct().order_by("curso"),
        "cursos_filtrados": Alumno.objects.filter(activo=True).exclude(curso="").values_list("curso", flat=True).distinct().order_by("curso"),
        "mes_actual": hoy.month,
        "anio_actual": hoy.year,
    })


@rol_requerido(INSPECTOR_GENERAL, PROFESOR, DIRECTOR)
def reporte_alumno(request, pk):
    alumno = get_object_or_404(Alumno, pk=pk)
    retiros = Retiro.objects.filter(alumno=alumno)
    atrs = Atraso.objects.filter(alumno=alumno)
    unifs = ControlUniforme.objects.filter(alumno=alumno)
    cels = Celular.objects.filter(alumno=alumno)

    fecha_desde = request.GET.get("fecha_desde", "")
    fecha_hasta = request.GET.get("fecha_hasta", "")
    if fecha_desde:
        retiros = retiros.filter(fecha__gte=fecha_desde)
        atrs = atrs.filter(fecha__gte=fecha_desde)
        unifs = unifs.filter(fecha__gte=fecha_desde)
        cels = cels.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        retiros = retiros.filter(fecha__lte=fecha_hasta)
        atrs = atrs.filter(fecha__lte=fecha_hasta)
        unifs = unifs.filter(fecha__lte=fecha_hasta)
        cels = cels.filter(fecha__lte=fecha_hasta)

    ctx = {
        "alumno": alumno,
        "retiros": retiros,
        "atrasos": atrs,
        "uniformes": unifs,
        "celulares": cels,
        "total_retiros": retiros.count(),
        "total_atrasos": atrs.count(),
        "total_uniformes": unifs.count(),
        "total_celulares": cels.count(),
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
    }
    return render(request, "core/reporte_alumno.html", ctx)


@rol_requerido(INSPECTOR_GENERAL, PROFESOR, DIRECTOR)
def reporte_curso(request, curso):
    alumnos_curso = Alumno.objects.filter(curso=curso, activo=True)
    fecha_desde = request.GET.get("fecha_desde", "")
    fecha_hasta = request.GET.get("fecha_hasta", "")
    datos = []
    for al in alumnos_curso:
        r_qs = al.retiros.all()
        a_qs = al.atrasos.all()
        u_qs = al.uniformes.all()
        c_qs = al.celulares.all()
        if fecha_desde:
            r_qs = r_qs.filter(fecha__gte=fecha_desde)
            a_qs = a_qs.filter(fecha__gte=fecha_desde)
            u_qs = u_qs.filter(fecha__gte=fecha_desde)
            c_qs = c_qs.filter(fecha__gte=fecha_desde)
        if fecha_hasta:
            r_qs = r_qs.filter(fecha__lte=fecha_hasta)
            a_qs = a_qs.filter(fecha__lte=fecha_hasta)
            u_qs = u_qs.filter(fecha__lte=fecha_hasta)
            c_qs = c_qs.filter(fecha__lte=fecha_hasta)
        datos.append({
            "alumno": al,
            "retiros": r_qs.count(),
            "atrasos": a_qs.count(),
            "uniformes": u_qs.count(),
            "celulares": c_qs.count(),
        })
    datos.sort(key=lambda x: x["atrasos"] + x["retiros"], reverse=True)
    return render(request, "core/reporte_curso.html", {
        "curso": curso, "datos": datos,
        "fecha_desde": fecha_desde, "fecha_hasta": fecha_hasta,
    })


def _metricas_curso(curso, mes, anio):
    return {
        "retiros": Retiro.objects.filter(alumno__curso=curso, fecha__month=mes, fecha__year=anio).count(),
        "atrasos": Atraso.objects.filter(alumno__curso=curso, fecha__month=mes, fecha__year=anio).count(),
        "uniformes": ControlUniforme.objects.filter(alumno__curso=curso, fecha__month=mes, fecha__year=anio).count(),
        "celulares": Celular.objects.filter(alumno__curso=curso, fecha__month=mes, fecha__year=anio).count(),
    }


def _metricas_alumno(alumno, mes, anio):
    return {
        "retiros": Retiro.objects.filter(alumno=alumno, fecha__month=mes, fecha__year=anio).count(),
        "atrasos": Atraso.objects.filter(alumno=alumno, fecha__month=mes, fecha__year=anio).count(),
        "uniformes": ControlUniforme.objects.filter(alumno=alumno, fecha__month=mes, fecha__year=anio).count(),
        "celulares": Celular.objects.filter(alumno=alumno, fecha__month=mes, fecha__year=anio).count(),
    }


def _metricas_curso_rango(curso, fd, fh):
    qs_r = Retiro.objects.filter(alumno__curso=curso)
    qs_a = Atraso.objects.filter(alumno__curso=curso)
    qs_u = ControlUniforme.objects.filter(alumno__curso=curso)
    qs_c = Celular.objects.filter(alumno__curso=curso)
    if fd:
        qs_r, qs_a, qs_u, qs_c = qs_r.filter(fecha__gte=fd), qs_a.filter(fecha__gte=fd), qs_u.filter(fecha__gte=fd), qs_c.filter(fecha__gte=fd)
    if fh:
        qs_r, qs_a, qs_u, qs_c = qs_r.filter(fecha__lte=fh), qs_a.filter(fecha__lte=fh), qs_u.filter(fecha__lte=fh), qs_c.filter(fecha__lte=fh)
    return {"retiros": qs_r.count(), "atrasos": qs_a.count(), "uniformes": qs_u.count(), "celulares": qs_c.count()}


def _metricas_alumno_rango(alumno, fd, fh):
    qs_r = Retiro.objects.filter(alumno=alumno)
    qs_a = Atraso.objects.filter(alumno=alumno)
    qs_u = ControlUniforme.objects.filter(alumno=alumno)
    qs_c = Celular.objects.filter(alumno=alumno)
    if fd:
        qs_r, qs_a, qs_u, qs_c = qs_r.filter(fecha__gte=fd), qs_a.filter(fecha__gte=fd), qs_u.filter(fecha__gte=fd), qs_c.filter(fecha__gte=fd)
    if fh:
        qs_r, qs_a, qs_u, qs_c = qs_r.filter(fecha__lte=fh), qs_a.filter(fecha__lte=fh), qs_u.filter(fecha__lte=fh), qs_c.filter(fecha__lte=fh)
    return {"retiros": qs_r.count(), "atrasos": qs_a.count(), "uniformes": qs_u.count(), "celulares": qs_c.count()}


@rol_requerido(INSPECTOR_GENERAL, DIRECTOR)
def reporte_general(request):
    from .pdf_generator import MESES
    hoy = date.today()
    mes = int(request.GET.get("mes", hoy.month))
    anio = int(request.GET.get("anio", hoy.year))
    if not 1 <= mes <= 12:
        mes = hoy.month
    curso = request.GET.get("curso", "").strip()
    fecha_desde = request.GET.get("fecha_desde", "")
    fecha_hasta = request.GET.get("fecha_hasta", "")
    usa_rango = bool(fecha_desde or fecha_hasta)

    cursos = list(
        Alumno.objects.filter(activo=True).exclude(curso="")
        .values_list("curso", flat=True).distinct().order_by("curso")
    )

    filas = []
    if curso and curso in cursos:
        alumnos_qs = Alumno.objects.filter(curso=curso, activo=True).order_by("apellido", "nombre")
        for al in alumnos_qs:
            if usa_rango:
                m = _metricas_alumno_rango(al, fecha_desde or None, fecha_hasta or None)
            else:
                m = _metricas_alumno(al, mes, anio)
            filas.append({"alumno": al, "nombre": al.nombre_completo, **m})
        filas.sort(key=lambda x: x["atrasos"] + x["retiros"], reverse=True)
    else:
        curso = ""
        for c in cursos:
            if usa_rango:
                m = _metricas_curso_rango(c, fecha_desde or None, fecha_hasta or None)
            else:
                m = _metricas_curso(c, mes, anio)
            filas.append({"alumno": None, "nombre": c, **m})
        filas.sort(key=lambda x: x["atrasos"] + x["retiros"] + x["uniformes"] + x["celulares"], reverse=True)

    totales = {
        k: sum(f[k] for f in filas)
        for k in ("retiros", "atrasos", "uniformes", "celulares")
    }
    totales["total"] = sum(totales.values())

    if usa_rango:
        rango_label = f"{fecha_desde or 'Inicio'} al {fecha_hasta or 'Hoy'}"
    else:
        rango_label = f"{MESES[mes]} {anio}"

    ctx = {
        "curso": curso,
        "cursos": cursos,
        "mes": mes,
        "anio": anio,
        "mes_nombre": rango_label,
        "filas": filas,
        "totales": totales,
        "detalle_por_alumno": bool(curso),
        "meses_opciones": list(MESES.items()),
        "anios_opciones": sorted({hoy.year, hoy.year - 1}, reverse=True),
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "usa_rango": usa_rango,
    }
    return render(request, "core/reporte_general.html", ctx)


@rol_requerido(INSPECTOR_GENERAL, DIRECTOR)
def exportar_excel_general(request):
    from .pdf_generator import MESES
    hoy = date.today()
    mes = int(request.GET.get("mes", hoy.month))
    anio = int(request.GET.get("anio", hoy.year))
    if not 1 <= mes <= 12:
        mes = hoy.month
    curso = request.GET.get("curso", "").strip()
    fecha_desde = request.GET.get("fecha_desde", "") or None
    fecha_hasta = request.GET.get("fecha_hasta", "") or None
    usa_rango = bool(fecha_desde or fecha_hasta)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen"
    ws.append(["Mes", "Alumno/Curso", "Retiros", "Atrasos", "Uniformes", "Celulares", "Total"])

    if curso:
        alumnos_qs = Alumno.objects.filter(curso=curso, activo=True).order_by("apellido", "nombre")
        for al in alumnos_qs:
            if usa_rango:
                m = _metricas_alumno_rango(al, fecha_desde, fecha_hasta)
            else:
                m = _metricas_alumno(al, mes, anio)
            ws.append([f"{MESES[mes]} {anio}" if not usa_rango else f"{fecha_desde or 'Inicio'} a {fecha_hasta or 'Hoy'}", al.nombre_completo, m["retiros"], m["atrasos"], m["uniformes"], m["celulares"],
                       m["retiros"] + m["atrasos"] + m["uniformes"] + m["celulares"]])
    else:
        cursos = (
            Alumno.objects.filter(activo=True).exclude(curso="")
            .values_list("curso", flat=True).distinct().order_by("curso")
        )
        for c in cursos:
            if usa_rango:
                m = _metricas_curso_rango(c, fecha_desde, fecha_hasta)
            else:
                m = _metricas_curso(c, mes, anio)
            ws.append([f"{MESES[mes]} {anio}" if not usa_rango else f"{fecha_desde or 'Inicio'} a {fecha_hasta or 'Hoy'}", c, m["retiros"], m["atrasos"], m["uniformes"], m["celulares"],
                       m["retiros"] + m["atrasos"] + m["uniformes"] + m["celulares"]])

    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="C8102E", end_color="C8102E", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for col, w in {"A": 14, "B": 34, "C": 10, "D": 10, "E": 10, "F": 10, "G": 10}.items():
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    label = curso.replace(" ", "_") if curso else "colegio"
    resp["Content-Disposition"] = f'attachment; filename="reporte_general_{label}_{mes}_{anio}.xlsx"'
    return resp


def _metricas_mes(mes, anio):
    return {
        "retiros": Retiro.objects.filter(fecha__month=mes, fecha__year=anio).count(),
        "atrasos": Atraso.objects.filter(fecha__month=mes, fecha__year=anio).count(),
        "uniformes": ControlUniforme.objects.filter(fecha__month=mes, fecha__year=anio).count(),
        "celulares": Celular.objects.filter(fecha__month=mes, fecha__year=anio).count(),
        "visitas": VisitaApoderado.objects.filter(fecha__month=mes, fecha__year=anio).count(),
    }


@rol_requerido(DIRECTOR)
def dashboard_director(request):
    from .pdf_generator import MESES
    hoy = date.today()
    mes = int(request.GET.get("mes", hoy.month))
    anio = int(request.GET.get("anio", hoy.year))
    if not 1 <= mes <= 12:
        mes = hoy.month

    if mes == 1:
        mes_ant, anio_ant = 12, anio - 1
    else:
        mes_ant, anio_ant = mes - 1, anio

    actual = _metricas_mes(mes, anio)
    anterior = _metricas_mes(mes_ant, anio_ant)

    comparacion = []
    for clave, label in [
        ("retiros", "Retiros"), ("atrasos", "Atrasos"),
        ("uniformes", "Uniformes"), ("celulares", "Celulares"),
        ("visitas", "Visitas"),
    ]:
        a, b = actual[clave], anterior[clave]
        delta = a - b
        pct = round(delta / b * 100) if b else (100 if a else 0)
        comparacion.append({
            "label": label,
            "actual": a,
            "anterior": b,
            "delta": delta,
            "pct": pct,
        })

    atrasos_por_curso = (
        Atraso.objects.filter(fecha__month=mes, fecha__year=anio)
        .values("alumno__curso")
        .annotate(total=Count("id"))
        .order_by("-total")
    )
    top_atrasos = (
        Atraso.objects.filter(fecha__month=mes, fecha__year=anio)
        .values("alumno__nombre", "alumno__apellido", "alumno__curso")
        .annotate(total=Count("id"))
        .order_by("-total")[:10]
    )
    retiros_por_motivo = (
        Retiro.objects.filter(fecha__month=mes, fecha__year=anio)
        .values("motivo")
        .annotate(total=Count("id"))
        .order_by("-total")
    )

    serie = []
    m, a = hoy.month, hoy.year
    for _ in range(6):
        serie.append({"mes": m, "anio": a, "etiqueta": f"{MESES[m][:3]} {str(a)[2:]}", **_metricas_mes(m, a)})
        if m == 1:
            m, a = 12, a - 1
        else:
            m -= 1
    serie.reverse()

    meses_disponibles = []
    m, a = hoy.month, hoy.year
    for _ in range(12):
        meses_disponibles.append({"mes": m, "anio": a})
        if m == 1:
            m, a = 12, a - 1
        else:
            m -= 1
    meses_disponibles.reverse()

    ctx = {
        "mes": mes,
        "anio": anio,
        "mes_ant": mes_ant,
        "anio_ant": anio_ant,
        "mes_nombre": f"{MESES[mes]} {anio}",
        "mes_ant_nombre": f"{MESES[mes_ant]} {anio_ant}",
        "actual": actual,
        "comparacion": comparacion,
        "atrasos_por_curso": atrasos_por_curso,
        "top_atrasos": top_atrasos,
        "retiros_por_motivo": retiros_por_motivo,
        "serie": serie,
        "meses_disponibles": meses_disponibles,
        "total_alumnos": Alumno.objects.filter(activo=True).count(),
    }
    return render(request, "core/dashboard_director.html", ctx)


# ── Llamadas a Apoderados (alumnos con 3+ atrasos, excluyendo Campo) ──
@rol_requerido(INSPECTOR_GENERAL, DIRECTOR)
def llamadas(request):
    form_llamada = LlamadaApoderadoForm()

    alumnos_queryset = (
        Alumno.objects.filter(activo=True)
        .annotate(
            atrasos_reales=Count("atrasos", filter=~Q(atrasos__motivo="CAMPO"))
        )
        .filter(atrasos_reales__gte=3)
        .order_by("-atrasos_reales", "apellido", "nombre")
    )

    alumnos_data = []
    for al in alumnos_queryset:
        llamadas_hist = al.llamadas.all()[:10]
        alumnos_data.append({
            "alumno": al,
            "total_atrasos": al.atrasos_reales,
            "llamadas": llamadas_hist,
        })

    if request.method == "POST" and "registrar_llamada" in request.POST:
        alumno_id = request.POST.get("alumno_id")
        alumno = get_object_or_404(Alumno, pk=alumno_id)
        form_llamada = LlamadaApoderadoForm(request.POST)
        if form_llamada.is_valid():
            llamada = form_llamada.save(commit=False)
            llamada.alumno = alumno
            llamada.registrado_por = request.user
            llamada.save()
            messages.success(request, f"Llamada registrada para {alumno.nombre_completo}.")
            return redirect("llamadas")

    return render(request, "core/llamadas.html", {
        "alumnos_data": alumnos_data,
        "form_llamada": form_llamada,
    })


# ── Gestión de usuarios (solo administrador) ──
@solo_admin
def usuarios(request):
    buscar = request.GET.get("buscar", "").strip()
    qs = User.objects.prefetch_related("groups").order_by("username")
    if buscar:
        qs = qs.filter(
            Q(username__icontains=buscar)
            | Q(first_name__icontains=buscar)
            | Q(last_name__icontains=buscar)
            | Q(email__icontains=buscar)
        )
    return render(request, "core/usuarios.html", {"usuarios": qs, "buscar": buscar})


@solo_admin
def usuario_crear(request):
    form = UsuarioCrearForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        user = form.save()
        messages.success(request, f"Usuario {user.username} creado correctamente.")
        return redirect("usuarios")
    return render(request, "core/usuario_form.html", {
        "form": form, "titulo": "Nuevo usuario",
    })


@solo_admin
def usuario_editar(request, pk):
    usuario = get_object_or_404(User, pk=pk)
    form = UsuarioForm(request.POST or None, instance=usuario)
    if request.method == "POST" and form.is_valid():
        if usuario == request.user:
            # Evita que el admin se bloqueque a sí mismo
            form.instance.is_active = True
            form.instance.is_superuser = True
        user = form.save()
        messages.success(request, f"Usuario {user.username} actualizado.")
        return redirect("usuarios")
    return render(request, "core/usuario_form.html", {
        "form": form, "titulo": f"Editar: {usuario.get_full_name() or usuario.username}",
        "usuario_obj": usuario,
    })


@solo_admin
def usuario_password(request, pk):
    usuario = get_object_or_404(User, pk=pk)
    form = UsuarioPasswordForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        usuario.set_password(form.cleaned_data["password1"])
        usuario.save()
        if usuario == request.user:
            update_session_auth_hash(request, usuario)
        messages.success(request, f"Contraseña restablecida para {usuario.username}.")
        return redirect("usuarios")
    return render(request, "core/usuario_password.html", {
        "form": form, "usuario_obj": usuario,
    })


# ── Ayuda / Manual de uso ──
MANUALES = [
    ("admin", "Administrador"),
    ("inspector_general", "Inspector General"),
    ("inspector", "Inspector"),
    ("profesor", "Profesor"),
    ("director", "Director"),
]


@login_required
def ayuda(request):
    if request.user.is_superuser:
        claves = [m[0] for m in MANUALES]
    elif tiene_rol(request.user, INSPECTOR_GENERAL):
        claves = ["inspector_general"]
    elif tiene_rol(request.user, INSPECTOR):
        claves = ["inspector"]
    elif tiene_rol(request.user, PROFESOR):
        claves = ["profesor"]
    elif tiene_rol(request.user, DIRECTOR):
        claves = ["director"]
    else:
        claves = []
    titulos = dict(MANUALES)
    secciones = [{"key": k, "titulo": titulos[k]} for k in claves]
    return render(request, "core/ayuda.html", {"secciones": secciones})


# ── Exportar PDF ──
@rol_requerido(INSPECTOR_GENERAL, PROFESOR, DIRECTOR)
def exportar_pdf_alumno(request, pk):
    from .pdf_generator import generar_pdf_alumno
    alumno = get_object_or_404(Alumno, pk=pk)
    buf = generar_pdf_alumno(alumno)
    resp = HttpResponse(buf, content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="reporte_{alumno.apellido}_{alumno.nombre}.pdf"'
    return resp


@rol_requerido(INSPECTOR_GENERAL, PROFESOR, DIRECTOR)
def exportar_pdf_curso(request, curso):
    from .pdf_generator import generar_pdf_curso
    mes = request.GET.get("mes")
    anio = request.GET.get("anio")
    mes = int(mes) if mes else None
    anio = int(anio) if anio else None
    fecha_desde = request.GET.get("fecha_desde") or None
    fecha_hasta = request.GET.get("fecha_hasta") or None
    buf = generar_pdf_curso(curso, mes=mes, anio=anio, fecha_desde=fecha_desde, fecha_hasta=fecha_hasta)
    resp = HttpResponse(buf, content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="informe_{curso.replace(" ","_")}.pdf"'
    return resp


@rol_requerido(INSPECTOR_GENERAL, PROFESOR, DIRECTOR)
def exportar_pdf_todos_cursos(request):
    from .pdf_generator import generar_pdf_todos_cursos
    mes = request.GET.get("mes")
    anio = request.GET.get("anio")
    mes = int(mes) if mes else None
    anio = int(anio) if anio else None
    fecha_desde = request.GET.get("fecha_desde") or None
    fecha_hasta = request.GET.get("fecha_hasta") or None
    buf = generar_pdf_todos_cursos(mes=mes, anio=anio, fecha_desde=fecha_desde, fecha_hasta=fecha_hasta)
    mes_label = mes or date.today().month
    anio_label = anio or date.today().year
    resp = HttpResponse(buf, content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="Informes_Todos_Cursos_{mes_label}_{anio_label}.pdf"'
    return resp


# ── Exportar Excel ──
@rol_requerido(INSPECTOR_GENERAL, PROFESOR, DIRECTOR)
def exportar_excel_alumno(request, pk):
    alumno = get_object_or_404(Alumno, pk=pk)
    wb = openpyxl.Workbook()

    # Retiros
    ws = wb.active
    ws.title = "Retiros"
    ws.append(["Fecha", "Hora", "Motivo", "Persona que retira", "RUT", "Observación"])
    for r in Retiro.objects.filter(alumno=alumno):
        ws.append([r.fecha, r.hora.strftime("%H:%M"), r.motivo, r.persona_retira, r.rut_retira, r.observacion])

    # Atrasos
    ws2 = wb.create_sheet("Atrasos")
    ws2.append(["Fecha", "Hora", "Tipo", "Lugar", "Observación"])
    for a in Atraso.objects.filter(alumno=alumno):
        ws2.append([a.fecha, a.hora.strftime("%H:%M"), a.tipo, a.lugar, a.observacion])

    # Uniformes
    ws3 = wb.create_sheet("Uniformes")
    ws3.append(["Fecha", "Falta", "Comprado", "Detalle", "Llamado", "Determinación"])
    for u in ControlUniforme.objects.filter(alumno=alumno):
        ws3.append([u.fecha, u.falta, "Sí" if u.tiene_uniforme_comprado else "No", u.detalle, "Sí" if u.llamado else "No", u.determinacion])

    # Celulares
    ws4 = wb.create_sheet("Celulares")
    ws4.append(["Fecha", "Lugar", "Retiro", "Aviso apoderado"])
    for c in Celular.objects.filter(alumno=alumno):
        ws4.append([c.fecha, c.lugar_entregado, c.retiro, "Sí" if c.aviso_apoderado else "No"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="reporte_{alumno.apellido}_{alumno.nombre}.xlsx"'
    return resp


@rol_requerido(INSPECTOR_GENERAL, PROFESOR, DIRECTOR)
def exportar_excel_curso(request, curso):
    alumnos_curso = Alumno.objects.filter(curso=curso, activo=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen"
    ws.append(["Alumno/a", "Retiros", "Atrasos", "Uniformes", "Celulares", "Total"])
    for al in alumnos_curso:
        r = al.retiros.count()
        a = al.atrasos.count()
        u = al.uniformes.count()
        c = al.celulares.count()
        ws.append([al.nombre_completo, r, a, u, c, r+a+u+c])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="reporte_{curso.replace(" ","_")}.xlsx"'
    return resp


# ── API: buscar alumnos (para autocompletar) ──
@login_required
def api_buscar_alumnos(request):
    from django.http import JsonResponse
    q = request.GET.get("q", "").strip()
    curso = request.GET.get("curso", "").strip()
    if len(q) < 2 and not curso:
        return JsonResponse([], safe=False)
    qs = Alumno.objects.filter(activo=True)
    if q:
        qs = qs.filter(Q(nombre__icontains=q) | Q(apellido__icontains=q))
    if curso:
        qs = qs.filter(curso=curso)
    qs = qs.order_by("apellido", "nombre")[:30]
    data = [{"id": a.id, "text": a.nombre_completo, "curso": a.curso, "es_campo": a.es_campo} for a in qs]
    return JsonResponse(data, safe=False)


# ── Cargar datos históricos desde Excel ──
@login_required
def cargar_historico(request):
    if not request.user.is_superuser:
        messages.error(request, "Solo administradores pueden cargar datos históricos.")
        return redirect("dashboard")

    if request.method == "POST" and request.FILES.get("archivo"):
        archivo = request.FILES["archivo"]
        try:
            from datetime import datetime as dt, time as t
            wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
            resultados = []
            errores = []  # lista de dicts: {hoja, fila, fecha, alumno, curso, motivo}

            def parse_date(val):
                if isinstance(val, date):
                    return val
                if hasattr(val, 'date'):
                    return val.date()
                if isinstance(val, str) and val:
                    try:
                        return dt.strptime(val, "%Y-%m-%d").date()
                    except ValueError:
                        return None
                return None

            def parse_time(val):
                if isinstance(val, t):
                    return val
                if hasattr(val, 'time'):
                    return val.time()
                if isinstance(val, str) and val:
                    try:
                        return dt.strptime(val, "%H:%M").time()
                    except ValueError:
                        return t(8, 0)
                return t(8, 0)

            def _es_verdadero(val):
                """Convierte SI/SÍ/True/1/YES a True"""
                if val is True or val == 1:
                    return True
                if val is None or val is False or val == 0:
                    return False
                return str(val).strip().upper() in ("SI", "SÍ", "YES", "TRUE", "1")

            def registrar_error(hoja, fila_num, fecha, alumno_txt, curso, motivo):
                errores.append({
                    "hoja": hoja,
                    "fila": fila_num,
                    "fecha": str(fecha) if fecha else "",
                    "alumno": alumno_txt,
                    "curso": curso,
                    "motivo": motivo,
                })

            # Cache de alumnos: pre-cargar todos los existentes en una sola query
            year = date.today().year
            alumnos_existentes = {}
            for al in Alumno.objects.filter(anio=year).only("id", "nombre", "apellido"):
                alumnos_existentes[(al.nombre, al.apellido)] = al
            alumnos_nuevos_buf = {}

            def get_alumno(nombre, apellido, curso):
                nombre = (nombre or "").strip().upper()
                apellido = (apellido or "").strip().upper()
                if not nombre or not apellido:
                    return None, nombre, apellido
                key = (nombre, apellido)
                if key in alumnos_existentes:
                    return alumnos_existentes[key], nombre, apellido
                if key in alumnos_nuevos_buf:
                    return alumnos_nuevos_buf[key], nombre, apellido
                al = Alumno(
                    nombre=nombre,
                    apellido=apellido,
                    anio=year,
                    curso=normalizar_curso(curso) or (curso or ""),
                )
                alumnos_nuevos_buf[key] = al
                return al, nombre, apellido

            def flush_alumnos():
                if alumnos_nuevos_buf:
                    Alumno.objects.bulk_create(list(alumnos_nuevos_buf.values()), ignore_conflicts=True)
                    for al in Alumno.objects.filter(
                        anio=year,
                        nombre__in=[n.nombre for n in alumnos_nuevos_buf.values()],
                        apellido__in=[n.apellido for n in alumnos_nuevos_buf.values()],
                    ):
                        alumnos_existentes[(al.nombre, al.apellido)] = al
                    alumnos_nuevos_buf.clear()

            # ── FASE 1: Pre-cargar alumnos y recolectar datos crudos ──
            raw_data = {}  # hoja -> lista de (idx, row_data_dict, al)

            if "Retiros" in wb.sheetnames:
                rows = []
                for idx, row in enumerate(wb["Retiros"].iter_rows(min_row=2, values_only=True), start=2):
                    rows.append((idx, row))
                raw_data["Retiros"] = rows

            if "Atrasos" in wb.sheetnames:
                rows = []
                for idx, row in enumerate(wb["Atrasos"].iter_rows(min_row=2, values_only=True), start=2):
                    rows.append((idx, row))
                raw_data["Atrasos"] = rows

            if "Uniformes" in wb.sheetnames:
                rows = []
                for idx, row in enumerate(wb["Uniformes"].iter_rows(min_row=2, values_only=True), start=2):
                    rows.append((idx, row))
                raw_data["Uniformes"] = rows

            if "Celulares" in wb.sheetnames:
                rows = []
                for idx, row in enumerate(wb["Celulares"].iter_rows(min_row=2, values_only=True), start=2):
                    rows.append((idx, row))
                raw_data["Celulares"] = rows

            if "Visitas" in wb.sheetnames:
                rows = []
                for idx, row in enumerate(wb["Visitas"].iter_rows(min_row=2, values_only=True), start=2):
                    rows.append((idx, row))
                raw_data["Visitas"] = rows

            # Descubrir todos los alumnos necesarios
            for hoja, rows in raw_data.items():
                if hoja == "Visitas":
                    continue
                for idx, row in rows:
                    if hoja == "Retiros":
                        get_alumno(row[1], row[2], row[3])
                    elif hoja == "Atrasos":
                        get_alumno(row[1], row[2], row[3])
                    elif hoja == "Uniformes":
                        get_alumno(row[1], row[2], row[3])
                    elif hoja == "Celulares":
                        get_alumno(row[1], row[2], row[3])

            # Flush: crear TODOS los alumnos nuevos de golpe
            flush_alumnos()

            # ── FASE 2: Crear registros usando alumnos ya guardados ──

            if "Retiros" in raw_data:
                objs = []
                for idx, row in raw_data["Retiros"]:
                    fecha = parse_date(row[0])
                    if not fecha:
                        registrar_error("Retiros", idx, row[0], f"{row[1]} {row[2]}", row[3], "Fecha inválida o vacía")
                        continue
                    if not row[1] or not row[2]:
                        registrar_error("Retiros", idx, fecha, f"{row[1]} {row[2]}", row[3], "Nombre o apellido vacío")
                        continue
                    key = ((row[1] or "").strip().upper(), (row[2] or "").strip().upper())
                    al = alumnos_existentes.get(key)
                    if not al:
                        registrar_error("Retiros", idx, fecha, f"{row[1]} {row[2]}", row[3], "No se pudo crear/encontrar alumno")
                        continue
                    objs.append(Retiro(
                        alumno=al, fecha=fecha, hora=parse_time(row[5]),
                        motivo=str(row[4] or "OTRO"),
                        persona_retira=str(row[6] or ""),
                        rut_retira=str(row[7] or ""),
                        registrado_por=request.user,
                    ))
                Retiro.objects.bulk_create(objs)
                resultados.append(("Retiros", len(objs), len([e for e in errores if e["hoja"] == "Retiros"])))

            if "Atrasos" in raw_data:
                objs = []
                for idx, row in raw_data["Atrasos"]:
                    fecha = parse_date(row[0])
                    if not fecha:
                        registrar_error("Atrasos", idx, row[0], f"{row[1]} {row[2]}", row[3], "Fecha inválida o vacía")
                        continue
                    if not row[1] or not row[2]:
                        registrar_error("Atrasos", idx, fecha, f"{row[1]} {row[2]}", row[3], "Nombre o apellido vacío")
                        continue
                    key = ((row[1] or "").strip().upper(), (row[2] or "").strip().upper())
                    al = alumnos_existentes.get(key)
                    if not al:
                        registrar_error("Atrasos", idx, fecha, f"{row[1]} {row[2]}", row[3], "No se pudo crear/encontrar alumno")
                        continue
                    objs.append(Atraso(
                        alumno=al, fecha=fecha, hora=parse_time(row[4]),
                        tipo=str(row[5] or "LLEGADA"),
                        lugar=str(row[6] or ""),
                        registrado_por=request.user,
                    ))
                Atraso.objects.bulk_create(objs)
                resultados.append(("Atrasos", len(objs), len([e for e in errores if e["hoja"] == "Atrasos"])))

            if "Uniformes" in raw_data:
                objs = []
                for idx, row in raw_data["Uniformes"]:
                    fecha = parse_date(row[0])
                    if not fecha:
                        registrar_error("Uniformes", idx, row[0], f"{row[1]} {row[2]}", row[3], "Fecha inválida o vacía")
                        continue
                    if not row[1] or not row[2]:
                        registrar_error("Uniformes", idx, fecha, f"{row[1]} {row[2]}", row[3], "Nombre o apellido vacío")
                        continue
                    key = ((row[1] or "").strip().upper(), (row[2] or "").strip().upper())
                    al = alumnos_existentes.get(key)
                    if not al:
                        registrar_error("Uniformes", idx, fecha, f"{row[1]} {row[2]}", row[3], "No se pudo crear/encontrar alumno")
                        continue
                    objs.append(ControlUniforme(
                        alumno=al, fecha=fecha,
                        falta=str(row[4] or "SIN UNIFORME"),
                        tiene_uniforme_comprado=_es_verdadero(row[5]),
                        detalle=str(row[6] or ""),
                        contacto_apoderado=str(row[7] or ""),
                        llamado=_es_verdadero(row[8]),
                        registrado_por=request.user,
                    ))
                ControlUniforme.objects.bulk_create(objs)
                resultados.append(("Uniformes", len(objs), len([e for e in errores if e["hoja"] == "Uniformes"])))

            if "Celulares" in raw_data:
                objs = []
                for idx, row in raw_data["Celulares"]:
                    fecha = parse_date(row[0])
                    if not fecha:
                        registrar_error("Celulares", idx, row[0], f"{row[1]} {row[2]}", row[3], "Fecha inválida o vacía")
                        continue
                    if not row[1] or not row[2]:
                        registrar_error("Celulares", idx, fecha, f"{row[1]} {row[2]}", row[3], "Nombre o apellido vacío")
                        continue
                    key = ((row[1] or "").strip().upper(), (row[2] or "").strip().upper())
                    al = alumnos_existentes.get(key)
                    if not al:
                        registrar_error("Celulares", idx, fecha, f"{row[1]} {row[2]}", row[3], "No se pudo crear/encontrar alumno")
                        continue
                    objs.append(Celular(
                        alumno=al, fecha=fecha,
                        lugar_entregado=str(row[4] or "DIRECCIÓN"),
                        retiro=str(row[5] or "AL FINAL DEL DÍA"),
                        aviso_apoderado=_es_verdadero(row[6]),
                        registrado_por=request.user,
                    ))
                Celular.objects.bulk_create(objs)
                resultados.append(("Celulares", len(objs), len([e for e in errores if e["hoja"] == "Celulares"])))

            if "Visitas" in raw_data:
                objs = []
                for idx, row in raw_data["Visitas"]:
                    fecha = parse_date(row[0])
                    if not fecha:
                        registrar_error("Visitas", idx, row[0], "", "", "Fecha inválida o vacía")
                        continue
                    objs.append(VisitaApoderado(
                        fecha=fecha, hora=parse_time(row[1]),
                        destino=str(row[2] or "INSPECTORÍA GENERAL"),
                        funcionario=str(row[3] or ""),
                        registrado_por=request.user,
                    ))
                VisitaApoderado.objects.bulk_create(objs)
                resultados.append(("Visitas", len(objs), len([e for e in errores if e["hoja"] == "Visitas"])))

            wb.close()

            # Guardar errores en sesión para descarga
            if errores:
                request.session["cargar_errores"] = errores
                request.session.modified = True

            total_ok = sum(r[1] for r in resultados)
            total_err = len(errores)
            ctx = {
                "resultados": resultados,
                "errores_count": total_err,
                "total_ok": total_ok,
                "tiene_errores": total_err > 0,
            }
            return render(request, "core/cargar_historico_resultado.html", ctx)

        except Exception as e:
            messages.error(request, f"Error al procesar el archivo: {e}")

    return render(request, "core/cargar_historico.html")


# ── Plantilla base para carga histórica ──
PLANTILLA_HISTORICO = {
    "Retiros": {
        "columnas": ["FECHA", "NOMBRE", "APELLIDO", "CURSO", "MOTIVO", "HORA", "PERSONA QUE RETIRA", "RUT QUIEN RETIRA"],
        "ejemplos": [
            ["2026-03-10", "PÉREZ", "JUAN", "1° BÁSICO", "ENFERMO/A", "10:30", "MARÍA ROJAS", "12345678-5"],
            ["2026-03-11", "GONZÁLEZ", "MARÍA", "2° BÁSICO", "TRÁMITES", "12:00", "PEDRO GONZÁLEZ", ""],
        ],
        "validaciones": {
            "E": "ENFERMO/A,CONTROL MÉDICO,PERSONAL,TRÁMITES,ACC. ESCOLAR,LOCOMOCIÓN,FAMILIAR,PERIODO ADAPTACIÓN,DEPORTE,ENTRENAMIENTO,ALMUERZO,VIAJE,TERAPIA,KINESIÓLOGO,OTRO",
        },
        "anchos": {"A": 12, "B": 18, "C": 18, "D": 14, "E": 20, "F": 10, "G": 26, "H": 16},
    },
    "Atrasos": {
        "columnas": ["FECHA", "NOMBRE", "APELLIDO", "CURSO", "HORA", "TIPO", "LUGAR"],
        "ejemplos": [
            ["2026-03-10", "PÉREZ", "JUAN", "1° BÁSICO", "08:15", "LLEGADA", "PORTERÍA"],
            ["2026-03-10", "GONZÁLEZ", "MARÍA", "2° BÁSICO", "10:35", "RECREO", "PATIO"],
        ],
        "validaciones": {"F": "LLEGADA,RECREO,ALMUERZO"},
        "anchos": {"A": 12, "B": 18, "C": 18, "D": 14, "E": 10, "F": 12, "G": 20},
    },
    "Uniformes": {
        "columnas": ["FECHA", "NOMBRE", "APELLIDO", "CURSO", "FALTA", "COMPRADO (SI/NO)", "DETALLE", "TELÉFONO CONTACTO", "LLAMADO (SI/NO)"],
        "ejemplos": [
            ["2026-03-12", "PÉREZ", "JUAN", "1° BÁSICO", "SIN POLERÓN", "SI", "No traía polerón", "912345678", "SI"],
            ["2026-03-12", "GONZÁLEZ", "MARÍA", "2° BÁSICO", "SIN UNIFORME", "NO", "", "", "NO"],
        ],
        "validaciones": {
            "E": "SIN UNIFORME,SIN POLERÓN,SIN POLERA NI POLERÓN",
            "F": "SI,NO",
            "I": "SI,NO",
        },
        "anchos": {"A": 12, "B": 18, "C": 18, "D": 14, "E": 24, "F": 16, "G": 28, "H": 18, "I": 16},
    },
    "Celulares": {
        "columnas": ["FECHA", "NOMBRE", "APELLIDO", "CURSO", "LUGAR ENTREGADO", "RETIRO", "AVISÓ APODERADO (SI/NO)"],
        "ejemplos": [
            ["2026-03-13", "PÉREZ", "JUAN", "1° BÁSICO", "DIRECCIÓN", "AL FINAL DEL DÍA", "SI"],
            ["2026-03-13", "GONZÁLEZ", "MARÍA", "2° BÁSICO", "INSPECTORÍA 1ER PISO", "RETIRA APODERADO", "NO"],
        ],
        "validaciones": {
            "E": "DIRECCIÓN,INSPECTORÍA 1ER PISO,INSPECTORÍA 2DO PISO",
            "F": "AL FINAL DEL DÍA,RETIRA APODERADO,PENDIENTE",
            "G": "SI,NO",
        },
        "anchos": {"A": 12, "B": 18, "C": 18, "D": 14, "E": 22, "F": 20, "G": 22},
    },
    "Visitas": {
        "columnas": ["FECHA", "HORA", "DESTINO", "FUNCIONARIO"],
        "ejemplos": [
            ["2026-03-14", "09:30", "UTP", "CARLA SOTO"],
            ["2026-03-14", "11:00", "INSPECTORÍA GENERAL", "LUIS HERRERA"],
        ],
        "validaciones": {
            "C": "INSPECTORÍA GENERAL,PIE,UTP,PROFESOR/A,CONV. ESCOLAR,DIRECCIÓN",
        },
        "anchos": {"A": 12, "B": 10, "C": 24, "D": 26},
    },
}


@login_required
def descargar_plantilla_historico(request):
    """Genera la planilla base .xlsx con hojas y columnas exactas que espera la carga histórica."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()

    # Hoja de instrucciones
    ws_info = wb.active
    ws_info.title = "Instrucciones"
    instrucciones = [
        "PLANILLA BASE — CARGA DE DATOS HISTÓRICOS",
        "",
        "1. No renombres ni elimines las hojas: Retiros, Atrasos, Uniformes, Celulares y Visitas.",
        "2. Mantén el orden exacto de las columnas de cada hoja.",
        "3. FECHA: formato AAAA-MM-DD (ej: 2026-03-15) o celda con formato de fecha.",
        "4. HORA: formato HH:MM en 24 horas (ej: 08:30). Si queda vacía se usa 08:00.",
        "5. NOMBRE y APELLIDO en MAYÚSCULAS, igual que en la planilla de alumnos.",
        "6. Campos SI/NO: escribe SI o NO.",
        "7. Usa los menús desplegables donde estén disponibles para evitar errores.",
        "8. ELIMINA las filas de ejemplo antes de subir el archivo.",
        "9. Las hojas son opcionales: si no tienes datos de una hoja, déjala solo con el encabezado.",
    ]
    for i, texto in enumerate(instrucciones, start=1):
        c = ws_info.cell(row=i, column=1, value=texto)
        if i == 1:
            c.font = Font(bold=True, size=13, color="C8102E")
    ws_info.column_dimensions["A"].width = 90

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="C8102E", end_color="C8102E", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    for nombre_hoja, cfg in PLANTILLA_HISTORICO.items():
        ws = wb.create_sheet(nombre_hoja)
        for col_idx, col_name in enumerate(cfg["columnas"], start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
        for ejemplo in cfg["ejemplos"]:
            ws.append(ejemplo)
        for col_letra, opciones in cfg["validaciones"].items():
            dv = DataValidation(type="list", formula1=f'"{opciones}"', allow_blank=True)
            dv.error = "Valor no permitido. Elige una opción de la lista."
            dv.errorTitle = "Dato inválido"
            ws.add_data_validation(dv)
            dv.add(f"{col_letra}2:{col_letra}1000")
        for col, w in cfg["anchos"].items():
            ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="planilla_base_carga_historica.xlsx"'
    return resp


@login_required
def descargar_errores_carga(request):
    if not request.user.is_superuser:
        messages.error(request, "Solo administradores.")
        return redirect("dashboard")
    errores = request.session.get("cargar_errores", [])
    if not errores:
        messages.warning(request, "No hay errores pendientes. Carga un archivo primero.")
        return redirect("cargar_historico")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Errores"
    headers = ["HOJA", "FILA", "FECHA", "ALUMNO", "CURSO", "MOTIVO_DEL_ERROR"]
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin = Side(style="thin", color="D9D9D9")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
        c.border = Border(bottom=thin)
    for r, err in enumerate(errores, start=2):
        ws.cell(row=r, column=1, value=err["hoja"])
        ws.cell(row=r, column=2, value=err["fila"])
        ws.cell(row=r, column=3, value=err["fecha"])
        ws.cell(row=r, column=4, value=err["alumno"])
        ws.cell(row=r, column=5, value=err["curso"])
        ws.cell(row=r, column=6, value=err["motivo"])
    for col, w in {"A": 14, "B": 8, "C": 14, "D": 32, "E": 16, "F": 36}.items():
        ws.column_dimensions[col].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = 'attachment; filename="errores_carga_historica.xlsx"'
    return resp


# ════════════════════════════════════════════
#  REPORTE DESDE EXCEL (sin guardar en BD)
# ════════════════════════════════════════════

EXCEL_COLUMNAS = ["FECHA", "APELLIDO_ALUMNO", "NOMBRE_ALUMNO", "CURSO", "HORA", "LLEGADA O RECREO", "LUGAR"]
TIPOS_VALIDOS = {"LLEGADA", "RECREO", "ALMUERZO"}


def _normalizar_header(s):
    import unicodedata
    s = (s or "").strip().upper()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return s


def _parse_date_excel(v):
    from datetime import datetime as dt
    if v is None or v == "":
        return None, "fecha vacía"
    if isinstance(v, dt):
        return v.date(), None
    if isinstance(v, date):
        return v, None
    if isinstance(v, (int, float)):
        try:
            from datetime import timedelta
            base = dt(1899, 12, 30)
            return (base + timedelta(days=int(v))).date(), None
        except Exception:
            return None, f"fecha numérica inválida ({v})"
    s = str(v).strip()
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%Y/%m/%d"):
        try:
            return dt.strptime(s, fmt).date(), None
        except ValueError:
            continue
    return None, f"fecha inválida '{s}'"


def _parse_time_excel(v):
    from datetime import datetime as dt, time as t
    if v is None or v == "":
        return None, "hora vacía"
    if isinstance(v, dt):
        return v.time(), None
    if isinstance(v, t):
        return v, None
    if isinstance(v, (int, float)):
        try:
            total_seconds = int(round(float(v) * 86400))
            hours = (total_seconds // 3600) % 24
            minutes = (total_seconds // 60) % 60
            seconds = total_seconds % 60
            return t(hours, minutes, seconds), None
        except Exception:
            return None, f"hora numérica inválida ({v})"
    s = str(v).strip()
    for fmt in ("%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M%p"):
        try:
            return dt.strptime(s, fmt).time(), None
        except ValueError:
            continue
    return None, f"hora inválida '{s}'"


def _normalizar_tipo(v):
    s = _normalizar_header(v)
    if s in TIPOS_VALIDOS:
        return s, None
    display = v if (v is not None and str(v).strip()) else "(vacío)"
    return s, f"tipo '{display}' no reconocido (se esperaba LLEGADA, RECREO o ALMUERZO)"


@rol_requerido(INSPECTOR_GENERAL, PROFESOR, DIRECTOR)
def reporte_desde_excel(request):
    """Muestra form de subida + procesa Excel y guarda filas en sesión para preview."""
    if request.method == "POST" and request.FILES.get("archivo"):
        archivo = request.FILES["archivo"]
        try:
            wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
            ws = wb.active
            filas = []
            headers = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [_normalizar_header(c) for c in row]
                    if not all(h in headers for h in EXCEL_COLUMNAS[:5]):
                        columnas_pos = EXCEL_COLUMNAS
                    continue
                if not row or all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
                    continue

                if headers and all(h in headers for h in EXCEL_COLUMNAS[:5]):
                    idx = {h: headers.index(h) for h in EXCEL_COLUMNAS}
                    get = lambda key, r=row: r[idx[key]] if idx[key] < len(r) else None
                else:
                    get = lambda key, ri=i - 1, r=row: r[ri] if ri < len(r) else None

                f_val, f_err = _parse_date_excel(get("FECHA"))
                a_val = (str(get("APELLIDO_ALUMNO") or "").strip().upper())
                n_val = (str(get("NOMBRE_ALUMNO") or "").strip().upper())
                c_val = (str(get("CURSO") or "").strip())
                h_val, h_err = _parse_time_excel(get("HORA"))
                t_val, t_err = _normalizar_tipo(get("LLEGADA O RECREO"))
                l_val = (str(get("LUGAR") or "").strip())

                errores = []
                if f_err:
                    errores.append(f_err)
                if not a_val:
                    errores.append("apellido vacío")
                if not n_val:
                    errores.append("nombre vacío")
                if h_err:
                    errores.append(h_err)
                if t_err:
                    errores.append(t_err)

                filas.append({
                    "fecha": f_val,
                    "apellido": a_val,
                    "nombre": n_val,
                    "curso": c_val,
                    "hora": h_val,
                    "tipo": t_val,
                    "lugar": l_val,
                    "errores": errores,
                })

            wb.close()

            from datetime import datetime as dt
            request.session["reporte_excel_meta"] = {
                "archivo_nombre": archivo.name,
                "generado": dt.now().strftime("%Y-%m-%d %H:%M:%S"),
                "total": len(filas),
                "con_errores": sum(1 for f in filas if f["errores"]),
                "validas": len(filas) - sum(1 for f in filas if f["errores"]),
            }
            session_filas = []
            for f in filas:
                session_filas.append({
                    "fecha": f["fecha"].isoformat() if f["fecha"] else "",
                    "apellido": f["apellido"],
                    "nombre": f["nombre"],
                    "curso": f["curso"],
                    "hora": f["hora"].strftime("%H:%M") if f["hora"] else "",
                    "tipo": f["tipo"],
                    "lugar": f["lugar"],
                    "errores": f["errores"],
                })
            request.session["reporte_excel_filas"] = session_filas
            request.session.modified = True

        except Exception as e:
            messages.error(request, f"Error al procesar el archivo: {e}")

    meta = request.session.get("reporte_excel_meta")
    filas_sesion = request.session.get("reporte_excel_filas", [])
    return render(request, "core/reporte_desde_excel.html", {
        "meta": meta,
        "filas": filas_sesion,
    })


@rol_requerido(INSPECTOR_GENERAL, PROFESOR, DIRECTOR)
def reporte_desde_excel_pdf(request):
    """Genera PDF desde las filas guardadas en sesión."""
    filas_sesion = request.session.get("reporte_excel_filas")
    meta = request.session.get("reporte_excel_meta")
    if not filas_sesion:
        messages.error(request, "No hay datos para generar el reporte. Sube un archivo primero.")
        return redirect("reporte_desde_excel")

    filas = []
    from datetime import date as d_cls, time as t_cls, datetime as dt_cls
    for f in filas_sesion:
        fecha = _parse_date_excel(f["fecha"])[0] if f["fecha"] else None
        hora = _parse_time_excel(f["hora"])[0] if f["hora"] else None
        filas.append({
            "fecha": fecha,
            "apellido": f["apellido"],
            "nombre": f["nombre"],
            "curso": f["curso"],
            "hora": hora,
            "tipo": f["tipo"],
            "lugar": f["lugar"],
            "errores": f.get("errores", []),
        })

    from .pdf_generator import generar_pdf_atrasos_excel
    buf = generar_pdf_atrasos_excel(filas, meta or {})
    nombre = (meta or {}).get("archivo_nombre", "reporte").rsplit(".", 1)[0]
    from datetime import datetime
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    resp = HttpResponse(buf, content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="Reporte_Atrasos_{ts}.pdf"'
    return resp


@rol_requerido(INSPECTOR_GENERAL, PROFESOR, DIRECTOR)
def descargar_plantilla_atrasos(request):
    """Genera un .xlsx de ejemplo con los encabezados de columnas esperados."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Atrasos"

    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="C8102E", end_color="C8102E", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    for col_idx, nombre in enumerate(EXCEL_COLUMNAS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=nombre)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    ws.append(["2026-03-15", "PÉREZ", "JUAN", "1° BÁSICO", "08:15", "LLEGADA", "Casa"])
    ws.append(["2026-03-15", "GONZÁLEZ", "MARÍA", "2° BÁSICO", "10:30", "RECREO", "Pasillo"])
    ws.append(["2026-03-16", "RAMÍREZ", "PEDRO", "I° MEDIO", "13:00", "ALMUERZO", "Comedor"])

    anchos = {"A": 12, "B": 18, "C": 18, "D": 14, "E": 10, "F": 18, "G": 20}
    for col, w in anchos.items():
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="plantilla_atrasos.xlsx"'
    return resp
